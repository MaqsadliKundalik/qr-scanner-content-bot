from aiogram import Router, F, Bot
from aiogram.types import Message, FSInputFile
from config import ADMINS
from filters.admin import IsAdminFilter
from keyboards.admin import admin_menu, admin_back_keyboard, hisobotlar_keyboard
from aiogram.fsm.context import FSMContext
from models.contents import QRCodes, Contents, Scans
from models.users import User
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
from datetime import datetime, timedelta

router = Router()

@router.message(F.text == "Kontentlar", IsAdminFilter())
async def contents_report(message: Message):
    contents = await Contents.all()
    contents_text = "Kontentlar hisobotı:\n\n"
    for content in contents:
        contents_text += f"/content{content.id} - {content.title}\nKontent turi: {content.content_type}\n\n"
        if len(contents_text) > 3900:
            await message.answer(contents_text)
            contents_text = ""
    if contents_text:
        total_contents = await Contents.all().count()
        contents_text += f"Jami kontentlar soni: {total_contents} ta"
        await message.answer(contents_text)

@router.message(F.text.startswith("/content"), IsAdminFilter())
async def content_detail_report(message: Message):
    content_id = message.text.replace("/content", "")
    selected_content = await Contents.get_or_none(id=content_id)
    if not selected_content:
        await message.answer("❌ Bunday kontent topilmadi.")
        return
    
    scan_count = await Scans.filter(content=selected_content).count()
    
    caption = selected_content.title if selected_content.title else ""
    caption += f"\n\nSkanerlashlar soni: {scan_count} ta"
    
    if selected_content.content_type == "text":
        await message.answer(selected_content.content)
    elif selected_content.content_type == "photo":
        await message.answer_photo(selected_content.content, caption=caption)
    elif selected_content.content_type == "video":
        await message.answer_video(selected_content.content, caption=caption)
    elif selected_content.content_type == "document":
        await message.answer_document(selected_content.content, caption=caption)
    
@router.message(F.text == "Skanerlashlar", IsAdminFilter())
async def scans_report(message: Message):
    users = await User.all().prefetch_related('scans')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Skanerlashlar hisoboti"
    
    ws['A1'] = "Skanerlashlar hisoboti"
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    headers = ['#', 'Telegram ID', 'Ism', 'Skanerlashlar soni']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row = 4
    total_scans = 0
    
    for idx, user in enumerate(users, start=1):
        scan_count = await Scans.filter(user=user).count()
        if scan_count > 0: 
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=user.telegram_id)
            ws.cell(row=row, column=3, value=user.name or "Noma'lum")
            ws.cell(row=row, column=4, value=scan_count)
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            
            total_scans += scan_count
            row += 1
    
    if row > 4:
        ws.cell(row=row, column=3, value="JAMI:")
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4, value=total_scans)
        ws.cell(row=row, column=4).font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    
    filename = "scans_report.xlsx"
    wb.save(filename)
    
    if row > 4:
        excel_file = FSInputFile(filename)
        await message.answer_document(
            excel_file,
            caption=f"📊 Skanerlashlar hisoboti\n\nJami foydalanuvchilar: {row - 4} ta\nJami skanerlashlar: {total_scans} ta"
        )
    else:
        await message.answer("❌ Hozircha hech kim skanerlamagan.")
    
    if os.path.exists(filename):
        os.remove(filename)

@router.message(F.text == "Foydalanuvchilar", IsAdminFilter())
async def users_report(message: Message):
    now = datetime.now()
    start_of_today = datetime(now.year, now.month, now.day)
    start_of_week = start_of_today - timedelta(days=start_of_today.weekday())
    start_of_month = datetime(now.year, now.month, 1)

    total_users = await User.all().count()
    users_today = await User.filter(created_at__gte=start_of_today).count()
    users_this_week = await User.filter(created_at__gte=start_of_week).count()
    users_this_month = await User.filter(created_at__gte=start_of_month).count()

    text = (
        f"👥 Foydalanuvchilar hisobotı:\n\n"
        f"🔹 Jami foydalanuvchilar: {total_users}\n"
        f"🔹 Shu oyda qo'shilganlar: {users_this_month}\n"
        f"🔹 Shu haftada qo'shilganlar: {users_this_week}\n"
        f"🔹 Bugun qo'shilganlar: {users_today}\n"
    )

    await message.answer(text)